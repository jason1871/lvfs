import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from config import (
    EXCEL_PATH, SHEET_DAILY, SHEET_MONTHLY,
    COLOR_HEADER_DAILY, COLOR_HEADER_MONTHLY, COLOR_TOTAL_ROW
)

DAILY_HEADERS = ['날짜', '중식', '석식', '합계', '비고']
MONTHLY_HEADERS = ['년월', '중식합계', '석식합계', '총합계', '중식평균', '석식평균', '급식일수']


def _thin_border():
    thin = Side(style='thin')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_style(fill_color):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    font = Font(bold=True, color='FFFFFFFF', size=11)
    align = Alignment(horizontal='center', vertical='center')
    return fill, font, align


def _apply_header_row(ws, headers, fill_color, row=1):
    fill, font, align = _header_style(fill_color)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = _thin_border()


def _set_column_widths(ws, widths):
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def get_or_create_workbook():
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
    else:
        wb = Workbook()
        # 기본 시트 제거
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # 일별집계 시트 생성
        ws_daily = wb.create_sheet(SHEET_DAILY)
        _apply_header_row(ws_daily, DAILY_HEADERS, COLOR_HEADER_DAILY)
        ws_daily.freeze_panes = 'A2'
        _set_column_widths(ws_daily, [14, 8, 8, 8, 20])

        # 월별집계 시트 생성
        ws_monthly = wb.create_sheet(SHEET_MONTHLY)
        _apply_header_row(ws_monthly, MONTHLY_HEADERS, COLOR_HEADER_MONTHLY)
        ws_monthly.freeze_panes = 'A2'
        _set_column_widths(ws_monthly, [10, 10, 10, 10, 10, 10, 10])

        wb.save(EXCEL_PATH)
    return wb


def _find_date_row(ws, date_str):
    """일별집계 시트에서 해당 날짜 행 번호 반환 (없으면 None)"""
    for row in ws.iter_rows(min_row=2, values_only=False):
        cell = row[0]
        val = cell.value
        if val is None:
            continue
        if isinstance(val, datetime):
            val = val.strftime('%Y-%m-%d')
        if str(val).strip() == date_str:
            return cell.row
    return None


def _style_data_row(ws, row_num, is_even=False):
    bg = 'FFF2F2F2' if is_even else 'FFFFFFFF'
    fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    for col in range(1, len(DAILY_HEADERS) + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.border = _thin_border()
        cell.fill = fill
        if col == 1:
            cell.alignment = align_center
        elif col == len(DAILY_HEADERS):
            cell.alignment = align_left
        else:
            cell.alignment = align_center


def append_daily(date_str, lunch, dinner):
    wb = get_or_create_workbook()
    ws = wb[SHEET_DAILY]

    existing_row = _find_date_row(ws, date_str)

    if existing_row:
        ws.cell(row=existing_row, column=2, value=lunch)
        ws.cell(row=existing_row, column=3, value=dinner)
        ws.cell(row=existing_row, column=4, value=lunch + dinner)
        print(f"  [{date_str}] 기존 행 업데이트 완료")
    else:
        next_row = ws.max_row + 1
        data = [date_str, lunch, dinner, lunch + dinner, '']
        for col, val in enumerate(data, 1):
            ws.cell(row=next_row, column=col, value=val)
        is_even = (next_row % 2 == 0)
        _style_data_row(ws, next_row, is_even)
        print(f"  [{date_str}] 새 행 추가 완료")

    update_monthly(wb)
    wb.save(EXCEL_PATH)
    print(f"  엑셀 저장: {EXCEL_PATH}")


def update_monthly(wb):
    ws_daily = wb[SHEET_DAILY]
    ws_monthly = wb[SHEET_MONTHLY]

    # 일별 데이터 읽기
    monthly_data = {}
    for row in ws_daily.iter_rows(min_row=2, values_only=True):
        date_val, lunch, dinner = row[0], row[1], row[2]
        if date_val is None or lunch is None:
            continue
        if isinstance(date_val, datetime):
            ym = date_val.strftime('%Y-%m')
        else:
            parts = str(date_val).strip().split('-')
            if len(parts) >= 2:
                ym = f"{parts[0]}-{parts[1]}"
            else:
                continue

        if ym not in monthly_data:
            monthly_data[ym] = {'중식': 0, '석식': 0, '일수': 0}
        monthly_data[ym]['중식'] += lunch or 0
        monthly_data[ym]['석식'] += dinner or 0
        monthly_data[ym]['일수'] += 1

    # 월별집계 시트 초기화 후 재작성
    for row in ws_monthly.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    fill_even = PatternFill(start_color='FFF2F2F2', end_color='FFF2F2F2', fill_type='solid')
    fill_odd = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')

    for i, (ym, data) in enumerate(sorted(monthly_data.items()), start=2):
        total = data['중식'] + data['석식']
        days = data['일수']
        lunch_avg = round(data['중식'] / days, 1) if days else 0
        dinner_avg = round(data['석식'] / days, 1) if days else 0

        row_data = [ym, data['중식'], data['석식'], total, lunch_avg, dinner_avg, days]
        fill = fill_even if i % 2 == 0 else fill_odd
        for col, val in enumerate(row_data, 1):
            cell = ws_monthly.cell(row=i, column=col, value=val)
            cell.border = _thin_border()
            cell.fill = fill
            cell.alignment = align_center

    # 합계 행
    last_data_row = ws_monthly.max_row
    if last_data_row >= 2:
        summary_row = last_data_row + 1
        total_lunch = sum(d['중식'] for d in monthly_data.values())
        total_dinner = sum(d['석식'] for d in monthly_data.values())
        total_days = sum(d['일수'] for d in monthly_data.values())
        total_all = total_lunch + total_dinner
        summary = ['합계', total_lunch, total_dinner, total_all, '', '', total_days]
        fill_total = PatternFill(start_color=COLOR_TOTAL_ROW, end_color=COLOR_TOTAL_ROW, fill_type='solid')
        font_bold = Font(bold=True)
        for col, val in enumerate(summary, 1):
            cell = ws_monthly.cell(row=summary_row, column=col, value=val)
            cell.border = _thin_border()
            cell.fill = fill_total
            cell.font = font_bold
            cell.alignment = align_center
