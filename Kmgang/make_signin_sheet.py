"""
인쇄용 식수명단 서식지 생성

파일명: {YY}년 {M}월 식수리스트.xlsx (예: 26년 5월 식수리스트.xlsx)
- 같은 달 실행 시 기존 파일에 날짜 시트 추가 (5.7, 5.8 ...)
- 월이 바뀌면 새 파일 자동 생성

사용법:
  python make_signin_sheet.py              # 오늘 날짜
  python make_signin_sheet.py 2026-05-07  # 특정 날짜
"""

import sys
import json
import os
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

MASTER_FILE = "master_members.json"

BRAND_COLORS = [
    "DDEEFF", "DFF2D8", "FFF2CC", "FCE4D6",
    "E2EFDA", "EDE7F6", "FDE8E8", "E8F5E9",
]

BLUE_DARK   = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
BLUE_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
WHITE_BOLD  = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(bold=True, size=14, color="FFFFFF")
CENTER      = Alignment(horizontal="center", vertical="center")


def _thin_border():
    thin = Side(style="thin", color="AAAAAA")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _monthly_filename(target_date: str) -> str:
    """'2026-05-07' → '26년 5월 식수리스트.xlsx'"""
    parts = target_date.split("-")
    yy = parts[0][2:]       # '2026' → '26'
    m  = str(int(parts[1])) # '05'   → '5'
    return f"{yy}년 {m}월 식수리스트.xlsx"


def _sheet_name(target_date: str) -> str:
    """'2026-05-07' → '5.7'"""
    parts = target_date.split("-")
    return f"{int(parts[1])}.{int(parts[2])}"


def load_master():
    with open(MASTER_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def _write_day_sheet(wb: openpyxl.Workbook, target_date: str, members: list) -> int:
    """워크북에 날짜 시트를 추가하고 행 수 반환."""
    sheet_name = _sheet_name(target_date)

    # 동일 날짜 시트가 이미 있으면 덮어쓰기
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # 브랜드 순서 유지
    brand_order = []
    for m in members:
        if m["brand"] not in brand_order:
            brand_order.append(m["brand"])

    # ── 행 1: 날짜 제목 ────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value     = f"아암센터 식수명단   {target_date}"
    c.font      = TITLE_FONT
    c.alignment = CENTER
    c.fill      = BLUE_DARK
    ws.row_dimensions[1].height = 28

    # ── 행 2: 컬럼 헤더 ───────────────────────────────────────────
    for col, label in enumerate(["브랜드", "이름", "중식", "석식"], 1):
        cell = ws.cell(row=2, column=col, value=label)
        cell.font      = WHITE_BOLD
        cell.fill      = BLUE_HEADER
        cell.alignment = CENTER
        cell.border    = _thin_border()
    ws.row_dimensions[2].height = 22

    # ── 행 3~N: 멤버 ─────────────────────────────────────────────
    current_row = 3
    for brand_idx, brand in enumerate(brand_order):
        fill = PatternFill(
            start_color=BRAND_COLORS[brand_idx % len(BRAND_COLORS)],
            end_color=BRAND_COLORS[brand_idx % len(BRAND_COLORS)],
            fill_type="solid",
        )
        for m in members:
            if m["brand"] != brand:
                continue
            ws.cell(row=current_row, column=1, value=m["brand"])
            ws.cell(row=current_row, column=2, value=m["name"])
            for col in range(1, 5):
                cell = ws.cell(row=current_row, column=col)
                cell.fill      = fill
                cell.border    = _thin_border()
                cell.alignment = CENTER
            ws.row_dimensions[current_row].height = 30
            current_row += 1

    # ── 열 너비 / 인쇄 설정 ───────────────────────────────────────
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth  = 1
    ws.print_area = f"A1:D{current_row - 1}"

    return current_row - 3  # 멤버 수


def make_sheet(target_date: str):
    out_path = _monthly_filename(target_date)
    members  = load_master()

    # 기존 월별 파일이 있으면 열고, 없으면 새로 생성
    if os.path.exists(out_path):
        wb = openpyxl.load_workbook(out_path)
        is_new = False
    else:
        wb = openpyxl.Workbook()
        # 기본 'Sheet' 제거
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        is_new = True

    member_count = _write_day_sheet(wb, target_date, members)
    wb.save(out_path)

    action = "신규 생성" if is_new else "시트 추가"
    sheet  = _sheet_name(target_date)
    print(f"[완료] {action}: {out_path}")
    print(f"  시트: {sheet} | 멤버 {member_count}명 | 브랜드 {len(set(m['brand'] for m in members))}개")
    print(f"  현재 시트 목록: {wb.sheetnames}")
    return out_path


def main():
    target_date = sys.argv[1] if len(sys.argv) > 1 else date.today().strftime("%Y-%m-%d")
    make_sheet(target_date)


if __name__ == "__main__":
    main()
