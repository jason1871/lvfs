"""
사용법:
  python process_meal.py '{"date":"2026-04-26","lunch":["이름1","이름2"],"dinner":["이름1"]}'
  python process_meal.py '{"date":"2026-04-26","lunch":["이름1"],"dinner":[]}'  # 저녁 없는 날
  python process_meal.py --close-month 2026-04  # 월 마감 (해당 월 별도 Excel 파일 저장)
"""
import sys
import json
import io
import re
from copy import copy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

EXCEL_FILE  = "아암센터 식수명단_26년 5월.xlsx"
MASTER_FILE = "master_members.json"

BLUE   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
WHITE  = Font(bold=True, color="FFFFFF")
CENTER = Alignment(horizontal="center")


def _stdout():
    return io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")


def _short_date(date_str: str) -> str:
    """'2026-05-04' → '5.4'"""
    parts = date_str.split("-")
    return f"{int(parts[1])}.{int(parts[2])}"


def load_master() -> list[dict]:
    with open(MASTER_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def _apply_header(ws, headers: list[str]) -> None:
    for col, header in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=header)
        c.font  = WHITE
        c.fill  = BLUE
        c.alignment = CENTER


def _add_brand_summary(ws, members: list[dict], start_col: int, lunch_col: int, dinner_col: int) -> None:
    """오른쪽 상단에 브랜드별 중식/석식 인원 요약표 추가."""
    ORANGE = PatternFill(start_color="F4B942", end_color="F4B942", fill_type="solid")

    # 브랜드 순서 유지
    seen = []
    for m in members:
        if m["brand"] not in seen:
            seen.append(m["brand"])
    if "영업" not in seen:
        seen.append("영업")
    brands = seen

    sc = start_col
    # 헤더
    for col_offset, label in enumerate(["브랜드", "중식", "석식"], sc):
        c = ws.cell(row=1, column=col_offset, value=label)
        c.font = WHITE
        c.fill = BLUE
        c.alignment = CENTER

    lunch_total_brand  = {b: 0 for b in brands}
    dinner_total_brand = {b: 0 for b in brands}

    # 마스터 멤버 행에서 집계
    for row in ws.iter_rows(min_row=2, values_only=False):
        brand_cell = row[0]
        lunch_cell = row[lunch_col - 1] if len(row) >= lunch_col else None
        dinner_cell = row[dinner_col - 1] if len(row) >= dinner_col else None
        brand = brand_cell.value
        if brand and brand in lunch_total_brand:
            if lunch_cell and lunch_cell.value:
                lunch_total_brand[brand] += 1
            if dinner_cell and dinner_cell.value:
                dinner_total_brand[brand] += 1

    for row_i, brand in enumerate(brands, 2):
        ws.cell(row=row_i, column=sc,     value=brand).alignment = CENTER
        ws.cell(row=row_i, column=sc + 1, value=lunch_total_brand[brand]).alignment  = CENTER
        ws.cell(row=row_i, column=sc + 2, value=dinner_total_brand[brand]).alignment = CENTER

    # 합계 행
    total_row = len(brands) + 2
    c = ws.cell(row=total_row, column=sc, value="합계")
    c.font = Font(bold=True)
    c.fill = ORANGE
    c.alignment = CENTER
    cl = ws.cell(row=total_row, column=sc + 1, value=sum(lunch_total_brand.values()))
    cl.font = Font(bold=True); cl.fill = ORANGE; cl.alignment = CENTER
    cd = ws.cell(row=total_row, column=sc + 2, value=sum(dinner_total_brand.values()))
    cd.font = Font(bold=True); cd.fill = ORANGE; cd.alignment = CENTER

    # 보장식수 행 (항상 표시, 값은 빈칸으로 사용자가 입력)
    BLUE_LIGHT = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    bojang_row = total_row + 1
    cb = ws.cell(row=bojang_row, column=sc, value="보장식수")
    cb.font = Font(bold=True); cb.fill = BLUE_LIGHT; cb.alignment = CENTER
    cbl = ws.cell(row=bojang_row, column=sc + 1, value=None)
    cbl.font = Font(bold=True); cbl.fill = BLUE_LIGHT; cbl.alignment = CENTER
    cbd = ws.cell(row=bojang_row, column=sc + 2, value=None)
    cbd.font = Font(bold=True); cbd.fill = BLUE_LIGHT; cbd.alignment = CENTER

    ws.column_dimensions[openpyxl.utils.get_column_letter(sc)].width     = 16
    ws.column_dimensions[openpyxl.utils.get_column_letter(sc + 1)].width = 8
    ws.column_dimensions[openpyxl.utils.get_column_letter(sc + 2)].width = 8


def _add_brand_summary_from_totals(ws, members: list[dict], lunch_total: dict, dinner_total: dict,
                                   extra_brand: dict, start_col: int) -> None:
    """집계 시트 오른쪽 상단에 브랜드별 누적 중식/석식 합계 요약표 추가.
    extra_brand: {이름 → 브랜드} — 마스터 외 인원, 해당 브랜드 버킷에 합산."""
    ORANGE = PatternFill(start_color="F4B942", end_color="F4B942", fill_type="solid")

    seen = []
    for m in members:
        if m["brand"] not in seen:
            seen.append(m["brand"])
    if "영업" not in seen:
        seen.append("영업")
    brands = seen

    brand_to_names = {}
    for m in members:
        brand_to_names.setdefault(m["brand"], []).append(m["name"])

    # 비마스터 인원을 브랜드 버킷에 추가
    for name, brand in extra_brand.items():
        if brand in brand_to_names:
            brand_to_names[brand].append(name)

    sc = start_col
    for col_offset, label in enumerate(["브랜드", "중식", "석식"], sc):
        c = ws.cell(row=1, column=col_offset, value=label)
        c.font = WHITE
        c.fill = BLUE
        c.alignment = CENTER

    for row_i, brand in enumerate(brands, 2):
        names = brand_to_names.get(brand, [])
        l_sum = sum(lunch_total.get(n, 0) for n in names)
        d_sum = sum(dinner_total.get(n, 0) for n in names)
        ws.cell(row=row_i, column=sc,     value=brand).alignment = CENTER
        ws.cell(row=row_i, column=sc + 1, value=l_sum).alignment  = CENTER
        ws.cell(row=row_i, column=sc + 2, value=d_sum).alignment  = CENTER

    total_row = len(brands) + 2
    c = ws.cell(row=total_row, column=sc, value="합계")
    c.font = Font(bold=True)
    c.fill = ORANGE
    c.alignment = CENTER
    # 합계는 마스터 + 비마스터 전체
    cl = ws.cell(row=total_row, column=sc + 1, value=sum(lunch_total.values()))
    cl.font = Font(bold=True); cl.fill = ORANGE; cl.alignment = CENTER
    cd = ws.cell(row=total_row, column=sc + 2, value=sum(dinner_total.values()))
    cd.font = Font(bold=True); cd.fill = ORANGE; cd.alignment = CENTER

    # 보장식수 행 (항상 표시)
    BLUE_LIGHT = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    bojang_row = total_row + 1
    cb = ws.cell(row=bojang_row, column=sc, value="보장식수")
    cb.font = Font(bold=True); cb.fill = BLUE_LIGHT; cb.alignment = CENTER
    cbl = ws.cell(row=bojang_row, column=sc + 1, value=None)
    cbl.font = Font(bold=True); cbl.fill = BLUE_LIGHT; cbl.alignment = CENTER
    cbd = ws.cell(row=bojang_row, column=sc + 2, value=None)
    cbd.font = Font(bold=True); cbd.fill = BLUE_LIGHT; cbd.alignment = CENTER

    ws.column_dimensions[openpyxl.utils.get_column_letter(sc)].width     = 16
    ws.column_dimensions[openpyxl.utils.get_column_letter(sc + 1)].width = 8
    ws.column_dimensions[openpyxl.utils.get_column_letter(sc + 2)].width = 8


def _read_bojang(date_sheets: list[str]) -> tuple:
    """각 날짜 시트 요약표(F열)에서 보장식수 행의 중식/석식 합계 읽기."""
    try:
        wb_r = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    except Exception:
        return 0, 0
    total_l = total_d = 0
    for ds in date_sheets:
        if ds not in wb_r.sheetnames:
            continue
        for row in wb_r[ds].iter_rows(min_col=6, values_only=True):
            if row[0] == "보장식수":
                if len(row) > 1 and isinstance(row[1], (int, float)):
                    total_l += int(row[1])
                if len(row) > 2 and isinstance(row[2], (int, float)):
                    total_d += int(row[2])
                break
    return total_l, total_d


def rebuild_monthly_summary(wb: openpyxl.Workbook, members: list[dict], ym: str) -> None:
    """해당 월의 모든 날짜 시트에서 집계 시트를 재계산 (덮어쓰기 재처리에도 정확)."""
    ORANGE = PatternFill(start_color="F4B942", end_color="F4B942", fill_type="solid")

    sheet_name  = f"{ym}_집계"
    date_sheets = [s for s in wb.sheetnames if re.match(r'^\d+\.\d+$', s)]

    master_names = {m["name"] for m in members}
    lunch_total  = {m["name"]: 0 for m in members}
    dinner_total = {m["name"]: 0 for m in members}
    # 비마스터: 브랜드별 누적
    extra_lunch  = {}  # brand → count
    extra_dinner = {}  # brand → count

    for ds in date_sheets:
        for row in wb[ds].iter_rows(min_row=2, values_only=True):
            brand = row[0] if len(row) > 0 else None
            name  = row[1] if len(row) > 1 else None
            if not name:
                continue
            if name in master_names:
                if len(row) > 2 and row[2]:
                    lunch_total[name] += 1
                if len(row) > 3 and row[3]:
                    dinner_total[name] += 1
            else:
                b = brand or "기타"
                if len(row) > 2 and row[2]:
                    extra_lunch[b]  = extra_lunch.get(b, 0) + 1
                if len(row) > 3 and row[3]:
                    extra_dinner[b] = extra_dinner.get(b, 0) + 1

    # 브랜드 순서 유지
    seen = []
    for m in members:
        if m["brand"] not in seen:
            seen.append(m["brand"])
    if "영업" not in seen:
        seen.append("영업")
    for b in list(extra_lunch) + list(extra_dinner):
        if b not in seen:
            seen.append(b)
    brands = seen

    brand_to_names = {}
    for m in members:
        brand_to_names.setdefault(m["brand"], []).append(m["name"])

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws_sum = wb.create_sheet(sheet_name)

    _apply_header(ws_sum, ["브랜드", "중식", "석식"])

    for row_i, brand in enumerate(brands, 2):
        names = brand_to_names.get(brand, [])
        l_sum = sum(lunch_total.get(n, 0) for n in names) + extra_lunch.get(brand, 0)
        d_sum = sum(dinner_total.get(n, 0) for n in names) + extra_dinner.get(brand, 0)
        ws_sum.cell(row=row_i, column=1, value=brand).alignment = CENTER
        ws_sum.cell(row=row_i, column=2, value=l_sum).alignment = CENTER
        ws_sum.cell(row=row_i, column=3, value=d_sum).alignment = CENTER

    # 보장식수 행
    bojang_lunch, bojang_dinner = _read_bojang(date_sheets)
    bojang_row = len(brands) + 2
    BLUE_LIGHT = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    cb = ws_sum.cell(row=bojang_row, column=1, value="보장식수")
    cb.font = Font(bold=True); cb.fill = BLUE_LIGHT; cb.alignment = CENTER
    cbl = ws_sum.cell(row=bojang_row, column=2, value=bojang_lunch)
    cbl.font = Font(bold=True); cbl.fill = BLUE_LIGHT; cbl.alignment = CENTER
    cbd = ws_sum.cell(row=bojang_row, column=3, value=bojang_dinner)
    cbd.font = Font(bold=True); cbd.fill = BLUE_LIGHT; cbd.alignment = CENTER

    total_row = bojang_row + 1
    total_lunch  = sum(lunch_total.values())  + sum(extra_lunch.values())
    total_dinner = sum(dinner_total.values()) + sum(extra_dinner.values())
    c = ws_sum.cell(row=total_row, column=1, value="합계")
    c.font = Font(bold=True); c.fill = ORANGE; c.alignment = CENTER
    cl = ws_sum.cell(row=total_row, column=2, value=total_lunch)
    cl.font = Font(bold=True); cl.fill = ORANGE; cl.alignment = CENTER
    cd = ws_sum.cell(row=total_row, column=3, value=total_dinner)
    cd.font = Font(bold=True); cd.fill = ORANGE; cd.alignment = CENTER

    ws_sum.column_dimensions["A"].width = 16
    ws_sum.column_dimensions["B"].width = 8
    ws_sum.column_dimensions["C"].width = 8


def _find_duplicates(name_list: list) -> set:
    """목록에서 중복 이름 반환"""
    seen = set()
    dupes = set()
    for n in name_list:
        if n in seen:
            dupes.add(n)
        seen.add(n)
    return dupes


def process(data: dict) -> None:
    date       = data["date"]
    lunch_raw  = data.get("lunch", [])
    dinner_raw = data.get("dinner", [])
    has_dinner = len(dinner_raw) > 0  # [] → 저녁 없는 날

    # 중복 이름 감지 (리스트 기준, set 변환 전)
    lunch_dupes  = _find_duplicates(lunch_raw)
    dinner_dupes = _find_duplicates(dinner_raw)
    dup_names    = lunch_dupes | dinner_dupes

    lunch_set  = set(lunch_raw)
    dinner_set = set(dinner_raw)

    members   = load_master()
    all_names = {m["name"] for m in members}

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet_name = _short_date(date)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    _apply_header(ws, ["브랜드", "이름", "중식", "석식"])

    for row_i, member in enumerate(members, 2):
        name = member["name"]

        lunch_val  = 1 if name in lunch_set else ""
        dinner_val = (1 if name in dinner_set else "") if has_dinner else ""

        ws.cell(row=row_i, column=1, value=member["brand"])
        name_cell = ws.cell(row=row_i, column=2, value=name)
        cl = ws.cell(row=row_i, column=3, value=lunch_val)
        cd = ws.cell(row=row_i, column=4, value=dinner_val)

        if name in dup_names:
            name_cell.fill = YELLOW  # 중복 이름 노랑 표시
        if lunch_val  == 1: cl.fill = GREEN
        if dinner_val == 1: cd.fill = GREEN

    # 마스터에 없는 이름 → 노랑 하이라이트 행으로 추가 (등록 문의 없이 표시만)
    brand_map = data.get("brand_map", {})
    unmatched = [n for n in (lunch_set | dinner_set) if n not in all_names]
    next_row  = len(members) + 2
    for name in unmatched:
        brand = brand_map.get(name, "?")
        ws.cell(row=next_row, column=1, value=brand)
        name_cell = ws.cell(row=next_row, column=2, value=name)
        name_cell.fill = YELLOW  # 마스터 불일치 또는 중복 → 노랑
        lc = ws.cell(row=next_row, column=3, value=1 if name in lunch_set  else "")
        dc = ws.cell(row=next_row, column=4, value=1 if name in dinner_set else "")
        if name in lunch_set:  lc.fill = GREEN
        if name in dinner_set: dc.fill = GREEN
        next_row += 1

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8

    _add_brand_summary(ws, members, start_col=6, lunch_col=3, dinner_col=4)

    rebuild_monthly_summary(wb, members, date[:7])

    wb.save(EXCEL_FILE)

    out = _stdout()
    out.write(f"[완료] [{sheet_name}] 시트 저장\n")
    out.write(f"  중식 {len(lunch_set)}명 | 석식 {'없음' if not has_dinner else f'{len(dinner_set)}명'}\n")
    if dup_names:
        out.write(f"[경고] 중복 이름 (노랑 표시): {', '.join(sorted(dup_names))}\n")
    if unmatched:
        out.write(f"[경고] 마스터 불일치 이름 (노랑 표시): {', '.join(unmatched)}\n")
    out.flush()


def close_month(ym: str) -> None:
    """월 마감: 해당 월의 모든 시트를 별도 Excel 파일로 저장."""
    out_file = f"{ym}_아암센터_식수명단.xlsx"

    wb_src    = openpyxl.load_workbook(EXCEL_FILE)
    target_sheets = [s for s in wb_src.sheetnames if s.startswith(ym)]

    if not target_sheets:
        out = _stdout()
        out.write(f"[경고] {ym}에 해당하는 시트가 없습니다.\n")
        out.flush()
        return

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    for sheet_name in target_sheets:
        ws_src = wb_src[sheet_name]
        ws_dst = wb_out.create_sheet(sheet_name)
        for row in ws_src.iter_rows():
            for cell in row:
                new_cell = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font      = copy(cell.font)
                    new_cell.fill      = copy(cell.fill)
                    new_cell.alignment = copy(cell.alignment)
        for col, dim in ws_src.column_dimensions.items():
            ws_dst.column_dimensions[col].width = dim.width

    wb_out.save(out_file)

    out = _stdout()
    out.write(f"[완료] 월 마감 파일 저장: {out_file}\n")
    out.write(f"  포함 시트 ({len(target_sheets)}개): {', '.join(target_sheets)}\n")
    out.flush()


if __name__ == "__main__":
    if len(sys.argv) >= 3 and sys.argv[1] == "--close-month":
        close_month(sys.argv[2])
    elif len(sys.argv) > 1:
        process(json.loads(sys.argv[1]))
    else:
        process(json.load(sys.stdin))
